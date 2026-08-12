"""Reading the structure out of a workbook a person built for other people.

A spreadsheet written by an analyst is not a table. It has a merged title banner, a header
split over two rows, a TOTAL line at the bottom, a blank column keeping two things apart,
and `N/A` where a number was not available. Every one of those is obvious to a reader and
invisible to a naive loader, which is why the reference implementation's would turn the
banner into a column name, the TOTAL row into a data row, and every merged cell but the
first into a NULL.

The profiler's job is to say what is actually there, so the ingest that follows has
something true to work from. It changes nothing and reads everything.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from vericlaim.sheets.profiler import profile_workbook


def workbook(tmp_path: Path, build) -> Path:
    book = Workbook()
    build(book)
    path = tmp_path / "Report.xlsx"
    book.save(path)
    return path


def plain(sheet) -> None:
    sheet.title = "Northern"
    sheet.append(["Region", "Claims", "Paid"])
    sheet.append(["Lahore", 12, 250000])
    sheet.append(["Karachi", 8, 180000])


def profile_one(tmp_path: Path, build):
    return profile_workbook(workbook(tmp_path, build))[0]


# ------------------------------------------------------------------ the simple case


def test_a_plain_sheet_has_one_table(tmp_path: Path) -> None:
    sheet = profile_one(tmp_path, lambda book: plain(book.active))

    assert len(sheet.tables) == 1
    assert sheet.sheet == "Northern"
    assert sheet.workbook == "Report.xlsx"


def test_the_header_names_the_columns(tmp_path: Path) -> None:
    table = profile_one(tmp_path, lambda book: plain(book.active)).tables[0]

    assert [column.label for column in table.columns] == ["Region", "Claims", "Paid"]
    assert [column.name for column in table.columns] == ["region", "claims", "paid"]


def test_the_data_rows_are_where_the_data_is(tmp_path: Path) -> None:
    table = profile_one(tmp_path, lambda book: plain(book.active)).tables[0]

    assert table.header_rows == (1,)
    assert (table.first_data_row, table.last_data_row) == (2, 3)


def test_the_table_knows_its_own_a1_range(tmp_path: Path) -> None:
    """Every citation this source will ever emit is built from this."""
    table = profile_one(tmp_path, lambda book: plain(book.active)).tables[0]

    assert table.a1_range == "A1:C3"


def test_a_column_knows_its_letter(tmp_path: Path) -> None:
    table = profile_one(tmp_path, lambda book: plain(book.active)).tables[0]

    assert [column.letter for column in table.columns] == ["A", "B", "C"]


# ------------------------------------------------------------------ the mess


def banner(book) -> None:
    sheet = book.active
    sheet.title = "Q1"
    sheet["A1"] = "Regional Inspection Compliance — Q1 2026"
    sheet.merge_cells("A1:C1")
    sheet.append([])
    sheet.append(["Region", "Inspections", "Compliance"])
    sheet.append(["Lahore", 40, 0.82])
    sheet.append(["Karachi", 35, 0.61])


def test_a_merged_title_banner_is_a_title_not_a_header(tmp_path: Path) -> None:
    """The reference loader would have made "Regional Inspection Compliance" a column."""
    table = profile_one(tmp_path, banner).tables[0]

    assert table.title == "Regional Inspection Compliance — Q1 2026"
    assert [column.label for column in table.columns] == [
        "Region",
        "Inspections",
        "Compliance",
    ]


def stacked(book) -> None:
    sheet = book.active
    sheet.title = "Targets"
    sheet["A1"] = "Region"
    sheet["B1"] = "Q1"
    sheet.merge_cells("B1:C1")
    sheet["A2"] = ""
    sheet["B2"] = "Target"
    sheet["C2"] = "Actual"
    sheet.append(["Lahore", 50, 42])
    sheet.append(["Karachi", 45, 47])


def test_a_two_row_header_becomes_one_name_per_column(tmp_path: Path) -> None:
    table = profile_one(tmp_path, stacked).tables[0]

    assert table.header_rows == (1, 2)
    assert [column.label for column in table.columns] == [
        "Region",
        "Q1 Target",
        "Q1 Actual",
    ]


def test_a_merged_header_cell_applies_across_the_columns_it_spans(tmp_path: Path) -> None:
    """Only the top-left cell holds the value; the rest read as empty. Taking that
    literally is what turns a merged header into two nameless columns."""
    table = profile_one(tmp_path, stacked).tables[0]

    assert table.columns[2].label == "Q1 Actual"


def totals(book) -> None:
    sheet = book.active
    sheet.title = "Loss"
    sheet.append(["Region", "Paid"])
    sheet.append(["Lahore", 250000])
    sheet.append(["Karachi", 180000])
    sheet.append(["TOTAL", 430000])


def test_a_total_line_is_a_footer_not_a_region(tmp_path: Path) -> None:
    """Left in the data, it doubles every aggregate and invents a region called TOTAL."""
    table = profile_one(tmp_path, totals).tables[0]

    assert table.footer_rows == (4,)
    assert table.last_data_row == 3


def two_tables(book) -> None:
    sheet = book.active
    sheet.title = "Both"
    sheet.append(["Region", "Claims"])
    sheet.append(["Lahore", 12])
    sheet.append([])
    sheet.append(["Adjuster", "Closed"])
    sheet.append(["A. Khan", 30])
    sheet.append(["S. Ali", 22])


def test_two_tables_on_one_sheet_stay_two_tables(tmp_path: Path) -> None:
    sheet = profile_one(tmp_path, two_tables)

    assert len(sheet.tables) == 2
    assert [column.label for column in sheet.tables[1].columns] == ["Adjuster", "Closed"]
    assert sheet.tables[1].first_data_row == 5


def spacer(book) -> None:
    sheet = book.active
    sheet.title = "Spaced"
    sheet["A1"], sheet["C1"], sheet["D1"] = "Region", "Claims", "Paid"
    sheet["A2"], sheet["C2"], sheet["D2"] = "Lahore", 12, 250000


def test_a_blank_spacer_column_is_not_a_column(tmp_path: Path) -> None:
    table = profile_one(tmp_path, spacer).tables[0]

    assert [column.letter for column in table.columns] == ["A", "C", "D"]
    assert table.spacer_columns == ("B",)


def sentinels(book) -> None:
    sheet = book.active
    sheet.title = "Gaps"
    sheet.append(["Region", "Compliance"])
    sheet.append(["Lahore", "N/A"])
    sheet.append(["Karachi", "-"])
    sheet.append(["Islamabad", 0.61])


def test_the_marks_people_use_for_missing_are_recorded(tmp_path: Path) -> None:
    """Coerced blindly they become the string "N/A" in a numeric column, and the column
    degrades to text for every row."""
    table = profile_one(tmp_path, sentinels).tables[0]

    assert set(table.columns[1].sentinels) == {"N/A", "-"}


def test_a_column_of_numbers_and_sentinels_is_still_a_number_column(tmp_path: Path) -> None:
    table = profile_one(tmp_path, sentinels).tables[0]

    assert table.columns[1].kind == "number"


# ------------------------------------------------------------------ formats


def formats(book) -> None:
    sheet = book.active
    sheet.title = "Money"
    sheet.append(["Region", "Paid", "Rate", "Reported"])
    sheet.append(["Lahore", 250000, 0.82, "2026-03-01"])
    sheet["B2"].number_format = '#,##0.00 "PKR"'
    sheet["C2"].number_format = "0.0%"
    sheet["D2"].number_format = "yyyy-mm-dd"


def test_a_currency_column_says_so(tmp_path: Path) -> None:
    """The number is 250000 either way; whether it is money decides how it may be added."""
    table = profile_one(tmp_path, formats).tables[0]

    assert table.columns[1].kind == "currency"


def test_a_percent_column_says_so(tmp_path: Path) -> None:
    """0.82 formatted as a percentage displays as 82%. A report quoting 0.82% is wrong by
    two orders of magnitude and looks entirely plausible."""
    table = profile_one(tmp_path, formats).tables[0]

    assert table.columns[2].kind == "percent"


def percent_as_text(book) -> None:
    sheet = book.active
    sheet.title = "Typed"
    sheet.append(["Region", "Rate"])
    sheet.append(["Lahore", "82%"])
    sheet.append(["Karachi", "61%"])


def test_a_percentage_typed_as_text_is_still_a_percentage(tmp_path: Path) -> None:
    table = profile_one(tmp_path, percent_as_text).tables[0]

    assert table.columns[1].kind == "percent"


def mixed(book) -> None:
    sheet = book.active
    sheet.title = "Mixed"
    sheet.append(["Region", "Claims"])
    sheet.append(["Lahore", 12])
    sheet.append(["Karachi", 8])
    sheet.append(["Islamabad", 5])
    sheet.append(["Multan", 3])
    sheet.append(["Quetta", "twelve"])


def test_one_stray_word_does_not_make_a_number_column_text(tmp_path: Path) -> None:
    """The reference's type inference degraded a whole column to TEXT on a single stray
    value, and every aggregate over it silently became impossible."""
    table = profile_one(tmp_path, mixed).tables[0]

    assert table.columns[1].kind == "number"
    assert table.columns[1].unparsed == ("twelve",)


# ------------------------------------------------------------------ edges


def test_an_empty_sheet_yields_no_tables(tmp_path: Path) -> None:
    def empty(book):
        book.active.title = "Blank"

    assert profile_one(tmp_path, empty).tables == ()


def test_every_sheet_in_the_workbook_is_profiled(tmp_path: Path) -> None:
    def several(book):
        plain(book.active)
        second = book.create_sheet("Southern")
        second.append(["Region", "Claims"])
        second.append(["Multan", 4])

    profiles = profile_workbook(workbook(tmp_path, several))

    assert [profile.sheet for profile in profiles] == ["Northern", "Southern"]


def test_a_workbook_that_is_not_there_says_so(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        profile_workbook(tmp_path / "missing.xlsx")


def test_a_merged_data_cell_keeps_its_value_across_the_span(tmp_path: Path) -> None:
    """openpyxl reports the value only on the top-left cell of a merge. Read literally,
    every other cell in the span becomes a NULL that was never in the spreadsheet."""

    def merged(book):
        sheet = book.active
        sheet.title = "Merged"
        sheet.append(["Region", "Q1", "Q2"])
        sheet["A2"] = "Lahore"
        sheet["B2"] = 12
        sheet.merge_cells("B2:C2")

    table = profile_one(tmp_path, merged).tables[0]

    assert table.value_at(2, 3) == 12
