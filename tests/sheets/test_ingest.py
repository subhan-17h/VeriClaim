"""Loading messy workbooks into Postgres without losing where anything came from.

The spreadsheets become `sheets.*` tables so the audited SQL validator and executor can
serve them, but they stay a distinct source: every ingested row carries the workbook, the
sheet, the row number and the A1 range it came from. Without those columns a query result
is a number nobody can check against the file, which is exactly the state the reference
implementation's loader left its data in.

Re-ingest is atomic. The reference dropped the table and then loaded, so a load that failed
half way left no data at all -- the state you are least able to recover from and most
likely to be in during a demo.
"""

from __future__ import annotations

from collections.abc import Iterator
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from vericlaim.config import Settings
from vericlaim.sheets.ingest import (
    LINEAGE_COLUMNS,
    generation_for,
    ingest_workbook,
    table_name,
)
from vericlaim.sheets.profiler import profile_workbook
from vericlaim.sql.db import Database

live = pytest.mark.postgres


def build(tmp_path: Path, name: str = "Regional_Report.xlsx") -> Path:
    book = Workbook()
    sheet = book.active
    sheet.title = "Q1 Compliance"
    sheet["A1"] = "Regional Inspection Compliance"
    sheet.merge_cells("A1:C1")
    sheet.append([])
    sheet.append(["Region", "Inspections", "Compliance"])
    sheet.append(["Lahore", 40, "82%"])
    sheet.append(["Karachi", 35, "N/A"])
    sheet.append(["TOTAL", 75, ""])
    path = tmp_path / name
    book.save(path)
    return path


# ------------------------------------------------------------------ naming


def test_a_table_is_named_for_the_workbook_and_the_sheet() -> None:
    """Two workbooks both with a "Summary" sheet must not collide, and the name has to be
    readable in a schema context a person reviews."""
    assert table_name("Regional_Report.xlsx", "Q1 Compliance", 0) == (
        "regional_report__q1_compliance"
    )


def test_a_second_table_on_one_sheet_gets_its_own_name() -> None:
    assert table_name("Report.xlsx", "Both", 1) == "report__both_2"


def test_a_name_that_would_start_with_a_digit_is_made_legal() -> None:
    assert table_name("2026_Report.xlsx", "Data", 0).startswith("_")


# ------------------------------------------------------------------ generation


def test_the_generation_is_the_content_of_the_workbook(tmp_path: Path) -> None:
    """Keyed on content, not on a timestamp: re-running the ingest over an unchanged file
    should be recognisable as the same load."""
    path = build(tmp_path)

    assert generation_for(path) == generation_for(path)


def test_editing_a_workbook_changes_its_generation(tmp_path: Path) -> None:
    first = generation_for(build(tmp_path))
    book = Workbook()
    book.active.append(["different"])
    book.save(tmp_path / "Regional_Report.xlsx")

    assert generation_for(tmp_path / "Regional_Report.xlsx") != first


# ------------------------------------------------------------------ live


@pytest.fixture
def admin(settings: Settings) -> Iterator[Database]:
    database = Database(settings.dsn(readonly=False), statement_timeout_ms=15_000)
    yield database
    with database.connection() as conn:
        conn.execute(
            "DO $$ DECLARE t text; BEGIN "
            "FOR t IN SELECT tablename FROM pg_tables WHERE schemaname = 'sheets' "
            "LOOP EXECUTE format('DROP TABLE IF EXISTS sheets.%I CASCADE', t); "
            "END LOOP; END $$;"
        )
    database.close()


def read(admin: Database, table: str) -> list[tuple]:
    with admin.connection() as conn:
        return conn.execute(
            f'SELECT * FROM sheets."{table}" ORDER BY _row'  # noqa: S608 - test-local
        ).fetchall()


@live
def test_a_workbook_becomes_a_table(tmp_path: Path, admin: Database) -> None:
    ingested = ingest_workbook(admin, build(tmp_path))

    assert len(ingested) == 1
    assert ingested[0].table == "regional_report__q1_compliance"
    assert ingested[0].row_count == 2


@live
def test_every_row_says_which_cells_it_came_from(tmp_path: Path, admin: Database) -> None:
    """This is the whole reason spreadsheets are a distinct source rather than more SQL."""
    ingested = ingest_workbook(admin, build(tmp_path))

    with admin.connection() as conn:
        rows = conn.execute(
            f'SELECT _workbook, _sheet, _row, _a1_range FROM sheets."{ingested[0].table}" '
            "ORDER BY _row"
        ).fetchall()

    assert rows[0] == ("Regional_Report.xlsx", "Q1 Compliance", 4, "A4:C4")


@live
def test_the_lineage_columns_are_on_every_table(tmp_path: Path, admin: Database) -> None:
    ingested = ingest_workbook(admin, build(tmp_path))

    with admin.connection() as conn:
        columns = {
            row[0]
            for row in conn.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = 'sheets' AND table_name = %s",
                (ingested[0].table,),
            ).fetchall()
        }

    assert set(LINEAGE_COLUMNS) <= columns


@live
def test_a_total_row_is_not_ingested_as_data(tmp_path: Path, admin: Database) -> None:
    """Ingested, it doubles every aggregate and invents a region called TOTAL."""
    ingested = ingest_workbook(admin, build(tmp_path))

    with admin.connection() as conn:
        regions = [
            row[0]
            for row in conn.execute(
                f'SELECT region FROM sheets."{ingested[0].table}"'
            ).fetchall()
        ]

    assert regions == ["Lahore", "Karachi"]


@live
def test_a_percentage_arrives_as_the_fraction_it_stands_for(
    tmp_path: Path, admin: Database
) -> None:
    ingested = ingest_workbook(admin, build(tmp_path))

    with admin.connection() as conn:
        value = conn.execute(
            f'SELECT compliance FROM sheets."{ingested[0].table}" WHERE region = %s',
            ("Lahore",),
        ).fetchone()[0]

    assert value == Decimal("0.82")


@live
def test_a_sentinel_arrives_as_null_not_as_the_word(
    tmp_path: Path, admin: Database
) -> None:
    ingested = ingest_workbook(admin, build(tmp_path))

    with admin.connection() as conn:
        value = conn.execute(
            f'SELECT compliance FROM sheets."{ingested[0].table}" WHERE region = %s',
            ("Karachi",),
        ).fetchone()[0]

    assert value is None


@live
def test_re_ingesting_replaces_the_rows_rather_than_doubling_them(
    tmp_path: Path, admin: Database
) -> None:
    path = build(tmp_path)
    ingest_workbook(admin, path)

    ingested = ingest_workbook(admin, path)

    assert ingested[0].row_count == 2
    assert len(read(admin, ingested[0].table)) == 2


@live
def test_a_failed_re_ingest_leaves_the_previous_data_in_place(
    tmp_path: Path, admin: Database
) -> None:
    """The reference dropped the table and then loaded. A load that failed half way left
    no data at all -- the state hardest to recover from and likeliest during a demo."""
    path = build(tmp_path)
    ingested = ingest_workbook(admin, path)
    profiles = profile_workbook(path)

    def explode(*args, **kwargs):
        raise RuntimeError("the load fell over")

    with pytest.raises(RuntimeError):
        ingest_workbook(admin, path, profiles=profiles, rows_for=explode)

    assert len(read(admin, ingested[0].table)) == 2


@live
def test_the_read_only_role_can_read_what_was_ingested(
    tmp_path: Path, admin: Database, settings: Settings
) -> None:
    """A table the agent cannot read is a table that does not exist as far as it knows."""
    ingested = ingest_workbook(admin, build(tmp_path))
    readonly = Database(
        settings.dsn(readonly=True), statement_timeout_ms=10_000, read_only_session=True
    )
    try:
        with readonly.connection() as conn:
            count = conn.execute(
                f'SELECT count(*) FROM sheets."{ingested[0].table}"'
            ).fetchone()[0]
    finally:
        readonly.close()

    assert count == 2


@live
def test_a_workbook_with_no_tables_ingests_nothing(
    tmp_path: Path, admin: Database
) -> None:
    book = Workbook()
    book.active.title = "Empty"
    path = tmp_path / "Empty.xlsx"
    book.save(path)

    assert ingest_workbook(admin, path) == ()
