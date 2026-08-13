"""The cross-source validator, and the corruptions it has to name.

A validator is only worth its runtime if it fails on a corpus that is actually wrong,
so every rule here is proved twice: once against a corpus built correctly, and once
against the same corpus with exactly one thing broken.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest

from vericlaim.config import Settings
from vericlaim.corpus.policies import generate_policy_corpus
from vericlaim.corpus.scanned import generate_scanned_corpus
from vericlaim.corpus.spreadsheets import generate_spreadsheet_corpus
from vericlaim.corpus.validate import Finding, validate_corpus

SEED = 42


@pytest.fixture(scope="module")
def corpus(tmp_path_factory: pytest.TempPathFactory) -> Settings:
    """Generate all three file sources once, into a settings pointing at them.

    ops.* is not loaded: the validator regenerates those rows in memory from the seed,
    which is what lets the whole check -- and this test -- run with no database.
    """
    root = tmp_path_factory.mktemp("corpus")
    settings = Settings().model_copy(
        update={
            "data_dir": root,
            "policy_dir": root / "policies",
            "spreadsheet_dir": root / "spreadsheets",
            "scanned_dir": root / "scanned",
        }
    )
    generate_policy_corpus(settings.policy_dir)
    generate_spreadsheet_corpus(settings.spreadsheet_dir, seed=SEED)
    generate_scanned_corpus(settings.scanned_dir, seed=SEED)
    return settings


def rules(findings: tuple[Finding, ...]) -> set[str]:
    return {finding.rule for finding in findings}


def test_a_correctly_generated_corpus_has_no_findings(corpus: Settings) -> None:
    assert validate_corpus(SEED, settings=corpus) == ()


def test_findings_are_ordered_the_same_way_on_every_run(corpus: Settings) -> None:
    first = validate_corpus(SEED, settings=corpus)
    second = validate_corpus(SEED, settings=corpus)

    assert first == second


def test_a_missing_policy_wording_is_named(corpus: Settings, tmp_path: Path) -> None:
    broken = _copy_of(corpus, tmp_path)
    target = next(broken.policy_dir.glob("HomeSecure*.pdf"))
    target.rename(target.with_name("Renamed_By_Accident.pdf"))

    findings = validate_corpus(SEED, settings=broken)

    assert "policy_document_missing" in rules(findings)
    assert any(target.name in finding.detail for finding in findings)


def test_a_scan_filed_under_an_unknown_claim_is_named(
    corpus: Settings, tmp_path: Path
) -> None:
    broken = _copy_of(corpus, tmp_path)
    scan = sorted(broken.scanned_dir.glob("*.pdf"))[0]
    scan.rename(scan.with_name("CLM-999999_INSPECTION_REPORT.pdf"))

    findings = validate_corpus(SEED, settings=broken)

    assert "scanned_claim_unknown" in rules(findings)
    assert any("CLM-999999" in finding.detail for finding in findings)


def test_a_scan_naming_no_claim_at_all_is_named(corpus: Settings, tmp_path: Path) -> None:
    broken = _copy_of(corpus, tmp_path)
    scan = sorted(broken.scanned_dir.glob("*.pdf"))[0]
    scan.rename(scan.with_name("loose_paperwork.pdf"))

    assert "scanned_filename_unkeyed" in rules(validate_corpus(SEED, settings=broken))


def test_an_undeclared_workbook_is_named(corpus: Settings, tmp_path: Path) -> None:
    broken = _copy_of(corpus, tmp_path)
    source = sorted(broken.spreadsheet_dir.glob("*.xlsx"))[0]
    shutil.copy2(source, broken.spreadsheet_dir / "Unreviewed_Extract.xlsx")

    findings = validate_corpus(SEED, settings=broken)

    assert "workbook_undeclared" in rules(findings)


def test_a_deleted_workbook_is_named(corpus: Settings, tmp_path: Path) -> None:
    broken = _copy_of(corpus, tmp_path)
    sorted(broken.spreadsheet_dir.glob("*.xlsx"))[0].unlink()

    findings = validate_corpus(SEED, settings=broken)

    assert "workbook_missing" in rules(findings)


def test_an_absent_source_directory_is_named_rather_than_passing_vacuously(
    corpus: Settings, tmp_path: Path
) -> None:
    """The most misleading result the validator could give is silence over nothing."""
    broken = _copy_of(corpus, tmp_path)
    shutil.rmtree(broken.scanned_dir)

    findings = validate_corpus(SEED, settings=broken)

    assert "source_incomplete" in rules(findings)
    assert any("scanned" in finding.detail for finding in findings)


def test_a_violated_invariant_is_named(corpus: Settings, tmp_path: Path) -> None:
    """More inspections completed than scheduled -- the ordered invariant the
    compliance context declares, checked over the coerced cells the ingest will read."""
    from openpyxl import load_workbook

    broken = _copy_of(corpus, tmp_path)
    path = next(broken.spreadsheet_dir.glob("Regional_Inspection_Compliance*.xlsx"))
    book = load_workbook(path)
    sheet = book["Compliance"]
    scheduled, completed = _columns(sheet, "scheduled", "completed")
    row = _first_data_row(sheet, scheduled)
    sheet.cell(row=row, column=completed).value = sheet.cell(row=row, column=scheduled).value + 5
    book.save(path)

    findings = validate_corpus(SEED, settings=broken)

    assert "sheet_invariant_violated" in rules(findings)


def test_a_violated_ops_invariant_is_named(corpus: Settings, monkeypatch) -> None:
    """The generated rows are judged by the same rule the observer applies to queries.

    Proved by breaking the arithmetic rather than the file: incurred is paid plus
    reserve, so a claim whose paid amount exceeds its incurred amount violates both the
    sum and the ordering ops.claims declares.
    """
    from dataclasses import replace
    from decimal import Decimal

    from vericlaim.corpus import validate as module

    real = module.generate_transactions

    def with_one_broken_claim(seed: int, **kwargs: object) -> object:
        rows = real(seed, **kwargs)  # type: ignore[arg-type]
        broken = replace(rows.claims[0], paid_amount_pkr=Decimal("999999999.00"))
        return replace(rows, claims=(broken, *rows.claims[1:]))

    monkeypatch.setattr(module, "generate_transactions", with_one_broken_claim)
    findings = validate_corpus(SEED, settings=corpus)

    assert "ops_invariant_violated" in rules(findings)
    assert any("ops.claims" in finding.detail for finding in findings)


def test_an_invariant_over_a_column_the_sheet_lacks_is_reported_not_skipped(
    corpus: Settings, tmp_path: Path
) -> None:
    """A rule that stops firing must not look like a rule that passed."""
    from openpyxl import load_workbook

    broken = _copy_of(corpus, tmp_path)
    path = next(broken.spreadsheet_dir.glob("Regional_Inspection_Compliance*.xlsx"))
    book = load_workbook(path)
    sheet = book["Compliance"]
    (scheduled,) = _columns(sheet, "scheduled")
    sheet.cell(row=3, column=scheduled).value = "planned_inspections"
    book.save(path)

    findings = validate_corpus(SEED, settings=broken)

    assert "sheet_column_missing" in rules(findings)


def _copy_of(corpus: Settings, tmp_path: Path) -> Settings:
    """A throwaway duplicate, so one test's corruption cannot reach another."""
    root = tmp_path / "corpus"
    shutil.copytree(corpus.data_dir, root)
    return corpus.model_copy(
        update={
            "data_dir": root,
            "policy_dir": root / "policies",
            "spreadsheet_dir": root / "spreadsheets",
            "scanned_dir": root / "scanned",
        }
    )


def _columns(sheet: object, *fragments: str) -> tuple[int, ...]:
    """Find each header column by a fragment of its label, whatever row it sits on."""
    found: list[int] = []
    for fragment in fragments:
        for row in sheet.iter_rows(min_row=1, max_row=12):  # type: ignore[attr-defined]
            match = next(
                (
                    cell.column
                    for cell in row
                    if isinstance(cell.value, str) and fragment in cell.value.lower()
                ),
                None,
            )
            if match is not None:
                found.append(match)
                break
        else:  # pragma: no cover - the fixture workbook always carries these headers
            raise AssertionError(f"no header column matching {fragment!r}")
    return tuple(found)


def _first_data_row(sheet: object, column: int) -> int:
    for row in range(1, 40):
        if isinstance(sheet.cell(row=row, column=column).value, int):  # type: ignore[attr-defined]
            return row
    raise AssertionError("no numeric data row found")  # pragma: no cover
