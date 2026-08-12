"""Read the structure out of a workbook a person built for other people.

A spreadsheet written by an analyst is not a table. It has a merged title banner, a header
split over two rows, a TOTAL line at the bottom, a blank column keeping two things apart,
and `N/A` where a number was not available. Every one of those is obvious to a human reader
and invisible to a loader that assumes row one is the header and everything below it is
data -- which is what the reference implementation assumed, and why it would turn the
banner into a column name, the TOTAL row into a region called TOTAL, and every merged cell
but the top-left one into a NULL that was never in the file.

This module changes nothing. It reads the sheet and says what is there: where each table
starts and ends, which rows are header and which are footer, what each column is called
once a stacked header is flattened, what kind of value it holds, and which marks were used
for "missing". The ingest in C-6.2 works from that description rather than from guesses,
and the A1 ranges recorded here are what every citation this source emits is built from.

Two deliberate judgements, both documented where they are made: a fully blank column splits
a block rather than being carried as a nameless column, and a column's kind is decided by
what most of its values are rather than by the first value that fails to parse.
"""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Marks people write when a number was not available. Kept as a set of *displayed* forms:
# a column of numbers with two of these in it is still a column of numbers.
NULL_SENTINELS = frozenset(
    {"n/a", "na", "n.a.", "-", "--", "—", "n/r", "nil", "tbd", "?", "#n/a"}
)
# A row whose first cell starts like this summarizes the rows above it.
FOOTER_RE = re.compile(
    r"^\s*(grand\s+)?(total|totals|subtotal|sum|average|mean|overall)\b", re.IGNORECASE
)
PERCENT_TEXT_RE = re.compile(r"^\s*-?[\d,]+(\.\d+)?\s*%\s*$")
NUMBER_TEXT_RE = re.compile(r"^\s*-?[\d,]+(\.\d+)?\s*$")
CURRENCY_FORMAT_RE = re.compile(r"(pkr|rs\.?|usd|[$£€₨¥])", re.IGNORECASE)
DATE_FORMAT_RE = re.compile(r"(yy|mm|dd|hh)", re.IGNORECASE)

# A column is called what most of its values are. One stray word in a column of numbers is
# a data-entry mistake, not a reason to give up on aggregating the column -- which is what
# the reference implementation did, degrading the whole column to text.
KIND_MAJORITY = 0.6


@dataclass(frozen=True, slots=True)
class ColumnProfile:
    """One column of one table: what it is called, and what it holds."""

    index: int
    letter: str
    label: str
    name: str
    kind: str
    sentinels: tuple[str, ...] = ()
    # Values that did not parse as the column's kind. Recorded rather than allowed to
    # change the kind, so the ingest can decide what to do about them and a reader can see
    # exactly which cells were a problem.
    unparsed: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class TableProfile:
    """One rectangular table found on a sheet, with its lineage."""

    title: str | None
    header_rows: tuple[int, ...]
    first_data_row: int
    last_data_row: int
    columns: tuple[ColumnProfile, ...]
    footer_rows: tuple[int, ...] = ()
    spacer_columns: tuple[str, ...] = ()
    values: dict[tuple[int, int], Any] | None = None

    @property
    def a1_range(self) -> str:
        first_column = self.columns[0].letter if self.columns else "A"
        last_column = self.columns[-1].letter if self.columns else "A"
        top = self.header_rows[0] if self.header_rows else self.first_data_row
        bottom = max([self.last_data_row, *self.footer_rows])
        return f"{first_column}{top}:{last_column}{bottom}"

    def row_range(self, row: int) -> str:
        """The A1 range of one data row, which is what a citation points at."""
        first_column = self.columns[0].letter if self.columns else "A"
        last_column = self.columns[-1].letter if self.columns else "A"
        return f"{first_column}{row}:{last_column}{row}"

    def value_at(self, row: int, column: int) -> Any:
        """The value a reader sees at a cell, with merges resolved."""
        return (self.values or {}).get((row, column))


@dataclass(frozen=True, slots=True)
class SheetProfile:
    """Every table found on one worksheet."""

    workbook: str
    sheet: str
    tables: tuple[TableProfile, ...] = ()
    merged_ranges: tuple[str, ...] = ()


def profile_workbook(path: Path | str) -> tuple[SheetProfile, ...]:
    """Profile every sheet of a workbook, in the order they appear in it."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"No such workbook: {path}")

    # data_only so formulas arrive as their last computed value: the profiler describes
    # what a reader sees, and "=SUM(B2:B9)" is not a value anyone can cite.
    book = load_workbook(path, data_only=True, read_only=False)
    try:
        return tuple(
            profile_sheet(book[name], workbook=path.name) for name in book.sheetnames
        )
    finally:
        book.close()


def profile_sheet(worksheet: Worksheet, *, workbook: str) -> SheetProfile:
    """Describe one worksheet: its merges, and every table on it."""
    merged = tuple(str(span) for span in worksheet.merged_cells.ranges)
    values = _resolved_values(worksheet)
    formats = _cell_formats(worksheet)
    tables = tuple(_tables(values, formats))
    return SheetProfile(
        workbook=workbook, sheet=worksheet.title, tables=tables, merged_ranges=merged
    )


# ------------------------------------------------------------------ cell values


def _resolved_values(worksheet: Worksheet) -> dict[tuple[int, int], Any]:
    """Every non-empty cell, with a merged span's value repeated across the span.

    openpyxl reports a merged value only on the top-left cell. Taking that literally is
    what turns "Q1" spanning two header columns into one named column and one nameless
    one, and a merged data cell into a NULL that was never in the spreadsheet.
    """
    values: dict[tuple[int, int], Any] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                values[(cell.row, cell.column)] = cell.value

    for span in worksheet.merged_cells.ranges:
        anchor = values.get((span.min_row, span.min_col))
        if anchor is None:
            continue
        for row in range(span.min_row, span.max_row + 1):
            for column in range(span.min_col, span.max_col + 1):
                values.setdefault((row, column), anchor)
    return values


# ------------------------------------------------------------------ finding tables


def _cell_formats(worksheet: Worksheet) -> dict[tuple[int, int], str]:
    """The display format of every non-empty cell.

    250000 is 250000 whether or not it is money, but whether it is money decides what it
    may be added to; and 0.82 displayed as 82% is wrong by two orders of magnitude when a
    report quotes the stored number instead of the shown one.
    """
    return {
        (cell.row, cell.column): cell.number_format or ""
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    }


def _tables(
    values: dict[tuple[int, int], Any], formats: dict[tuple[int, int], str]
) -> Iterator[TableProfile]:
    """Split the sheet into blocks of adjacent non-empty rows, and profile each.

    A block that is nothing but a banner -- one value across the width, with no data under
    it -- is not a table. It is the title of the block that follows, which is how an
    analyst writes a heading: on its own line, with a blank row under it.
    """
    occupied = sorted({row for row, _ in values})
    pending_title: str | None = None
    for block in _blocks(occupied):
        banner = _banner(values, block)
        if banner is not None:
            pending_title = banner
            continue
        table = _table(values, formats, block, pending_title)
        pending_title = None
        if table is not None:
            yield table


def _banner(values: dict[tuple[int, int], Any], block: list[int]) -> str | None:
    """Return the heading text if the block is only a heading."""
    if len(block) != 1:
        return None
    present = [value for (row, _), value in values.items() if row == block[0]]
    if not present or len({str(value) for value in present}) != 1:
        return None
    return str(present[0]).strip()


def _blocks(rows: Sequence[int]) -> Iterator[list[int]]:
    """Group row numbers into runs separated by at least one blank row."""
    block: list[int] = []
    for row in rows:
        if block and row != block[-1] + 1:
            yield block
            block = []
        block.append(row)
    if block:
        yield block


def _table(
    values: dict[tuple[int, int], Any],
    formats: dict[tuple[int, int], str],
    block: list[int],
    pending_title: str | None = None,
) -> TableProfile | None:
    columns = sorted({column for row, column in values if row in block})
    if not columns:
        return None

    inline_title, block = _title(values, block, columns)
    title = inline_title or pending_title
    if not block:
        return None

    header_rows = _header_rows(values, block, columns)
    if not header_rows:
        return None

    data_rows = [row for row in block if row > header_rows[-1]]
    footer_rows = tuple(_footer_rows(values, data_rows, columns))
    data_rows = [row for row in data_rows if row not in footer_rows]
    if not data_rows:
        return None

    used, spacers = _used_columns(values, data_rows, columns)
    profiles = tuple(
        _column(values, formats, header_rows, data_rows, index, column)
        for index, column in enumerate(used)
    )
    return TableProfile(
        title=title,
        header_rows=tuple(header_rows),
        first_data_row=data_rows[0],
        last_data_row=data_rows[-1],
        columns=profiles,
        footer_rows=footer_rows,
        spacer_columns=tuple(get_column_letter(column) for column in spacers),
        values=values,
    )


def _title(
    values: dict[tuple[int, int], Any], block: list[int], columns: list[int]
) -> tuple[str | None, list[int]]:
    """Peel a one-value banner row off the top of a block.

    A row holding a single distinct value across the width of the table is a heading
    somebody wrote for a human. Left in place it becomes the header row, and the real
    header becomes the first row of data.
    """
    if len(block) < 2:
        return None, block
    first = block[0]
    row_values = [values.get((first, column)) for column in columns]
    present = [value for value in row_values if value is not None]
    if len(present) >= 1 and len({str(value) for value in present}) == 1:
        below = [values.get((block[1], column)) for column in columns]
        if sum(value is not None for value in below) > len(present):
            return str(present[0]).strip(), block[1:]
    return None, block


def _header_rows(
    values: dict[tuple[int, int], Any], block: list[int], columns: list[int]
) -> list[int]:
    """Return the one or two rows that name the columns.

    A header row is text throughout. A second row is part of the header when it is also
    text throughout and the row below it is not -- which is what a stacked header looks
    like: "Q1" spanning two columns above "Target" and "Actual".
    """
    first = block[0]
    if not _is_text_row(values, first, columns):
        return []
    if len(block) >= 3 and _is_text_row(values, block[1], columns):
        if not _is_text_row(values, block[2], columns):
            return [first, block[1]]
    return [first]


def _is_text_row(
    values: dict[tuple[int, int], Any], row: int, columns: list[int]
) -> bool:
    present = [values.get((row, column)) for column in columns]
    present = [value for value in present if value is not None]
    if not present:
        return False
    return all(isinstance(value, str) for value in present)


def _footer_rows(
    values: dict[tuple[int, int], Any], rows: Sequence[int], columns: Sequence[int]
) -> Iterator[int]:
    """Trailing rows that summarize the ones above them.

    Only trailing rows: a row saying "Total" in the middle of a sheet is a subtotal for a
    group, and dropping it silently would remove data the rest of the table depends on.
    Left in the data, a footer doubles every aggregate and invents an entity called TOTAL.
    """
    for row in reversed(rows):
        first = next(
            (values.get((row, column)) for column in columns if (row, column) in values),
            None,
        )
        if isinstance(first, str) and FOOTER_RE.match(first):
            yield row
        else:
            return


def _used_columns(
    values: dict[tuple[int, int], Any], rows: Sequence[int], columns: list[int]
) -> tuple[list[int], list[int]]:
    """Split the columns into those holding data and the blank ones between them.

    A fully blank column is a spacer somebody added for legibility. It carries no data
    either way, so dropping it costs nothing and keeps it out of the ingested schema as a
    nameless column of NULLs.
    """
    span = range(columns[0], columns[-1] + 1)
    used = [column for column in span if any((row, column) in values for row in rows)]
    spacers = [column for column in span if column not in used]
    return used, spacers


# ------------------------------------------------------------------ columns


def _column(
    values: dict[tuple[int, int], Any],
    formats: dict[tuple[int, int], str],
    header_rows: Sequence[int],
    data_rows: Sequence[int],
    index: int,
    column: int,
) -> ColumnProfile:
    parts = [
        str(values[(row, column)]).strip()
        for row in header_rows
        if (row, column) in values
    ]
    # A stacked header repeats the merged upper label on every column it spans, so the two
    # parts join into a name that is unique across the table: "Q1 Target", "Q1 Actual".
    label = " ".join(dict.fromkeys(part for part in parts if part))
    cells = [values.get((row, column)) for row in data_rows]
    kind, sentinels, unparsed = _kind(cells)
    kind = _refine(kind, [formats.get((row, column), "") for row in data_rows])
    return ColumnProfile(
        index=index,
        letter=get_column_letter(column),
        label=label or get_column_letter(column),
        name=normalize_name(label, fallback=f"column_{index + 1}"),
        kind=kind,
        sentinels=sentinels,
        unparsed=unparsed,
    )


def _kind(cells: Sequence[Any]) -> tuple[str, tuple[str, ...], tuple[str, ...]]:
    """Decide what a column holds, and record what did not fit.

    Sentinels are set aside before anything is counted: `N/A` is an absent number, not a
    string, and letting it vote would make every column with a gap in it a text column.
    """
    sentinels: list[str] = []
    considered: list[Any] = []
    for cell in cells:
        if isinstance(cell, str) and cell.strip().lower() in NULL_SENTINELS:
            sentinels.append(cell.strip())
        elif cell is not None:
            considered.append(cell)

    if not considered:
        return "empty", tuple(dict.fromkeys(sentinels)), ()

    votes = [_cell_kind(cell) for cell in considered]
    winner = max(sorted(set(votes)), key=votes.count)
    if votes.count(winner) < KIND_MAJORITY * len(votes):
        winner = "text"

    unparsed = tuple(
        str(cell)
        for cell, vote in zip(considered, votes, strict=True)
        if vote != winner and winner != "text"
    )
    return winner, tuple(dict.fromkeys(sentinels)), unparsed


def _cell_kind(cell: Any) -> str:
    if isinstance(cell, bool):
        return "boolean"
    if isinstance(cell, (datetime, date)):
        return "date"
    if isinstance(cell, (int, float)):
        return "number"
    text = str(cell).strip()
    if PERCENT_TEXT_RE.match(text):
        return "percent"
    if NUMBER_TEXT_RE.match(text):
        return "number"
    return "text"


def _refine(kind: str, number_formats: Sequence[str]) -> str:
    """Narrow a plain number column using the format its cells are displayed in.

    Only a number can be narrowed: a text column formatted as currency is still text, and
    a column that is already a percentage or a date needs no help.
    """
    if kind != "number":
        return kind
    for number_format in number_formats:
        if "%" in number_format:
            return "percent"
        if CURRENCY_FORMAT_RE.search(number_format):
            return "currency"
        if DATE_FORMAT_RE.search(number_format):
            return "date"
    return kind


def normalize_name(value: object, fallback: str = "column") -> str:
    """Fold a header label into an identifier a database can hold."""
    text = unicodedata.normalize("NFKD", str(value))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    if not text:
        return fallback
    if text[0].isdigit():
        return f"_{text}"
    return text
