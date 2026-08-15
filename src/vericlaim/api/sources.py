"""Serving the source a piece of evidence came from.

Read-only, and deliberately narrow: every route answers "show me where this evidence
came from" and none answers "show me everything you have". There is no corpus to
browse without a question.

A name that arrives from a client is looked up, never joined. The catalog is built
from what exists, so `../../etc/passwd` is simply a name it does not hold -- which is
stronger than sanitising the input, because it cannot be defeated by an encoding the
sanitiser did not anticipate.
"""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path

from fastapi import APIRouter, HTTPException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from starlette.responses import FileResponse

from vericlaim.config import Settings
from vericlaim.sql.contexts import SchemaContext, context_detail, load_contexts

#: The largest sheet this will render. An order of magnitude above the largest sheet in
#: this corpus, so it bounds a pathological workbook without silently truncating a real
#: one -- and a truncated grid says so either way.
MAX_SHEET_ROWS = 500


def _pdfs(directory: Path) -> dict[str, Path]:
    """The PDFs in a directory, by name. Absent directory means an empty corpus."""
    if not directory.is_dir():
        return {}
    return {path.name: path for path in sorted(directory.glob("*.pdf"))}


def _cell(value: object) -> str:
    """Render a cell as the sheet holds it.

    Stringified here rather than in the client so nothing downstream reformats a
    number into something the sheet does not say. An empty cell is an empty string,
    not a null: the grid is a rectangle.
    """
    if value is None:
        return ""
    return str(value)


def read_sheet(
    path: Path, sheet: str, *, limit: int = MAX_SHEET_ROWS
) -> dict[str, object]:
    """Read one sheet as written: banner, blank rows, header and all."""
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise KeyError(sheet)
        worksheet = book[sheet]
        width = worksheet.max_column or 0
        rows: list[list[str]] = []
        total = 0
        for values in worksheet.iter_rows(values_only=True):
            total += 1
            if len(rows) < limit:
                padded = list(values) + [None] * (width - len(values))
                rows.append([_cell(value) for value in padded[:width]])
        return {
            "workbook": path.name,
            "sheet": sheet,
            "columns": [get_column_letter(index + 1) for index in range(width)],
            "first_row": 1,
            "rows": rows,
            "total_rows": total,
            "truncated": total > len(rows),
        }
    finally:
        book.close()


@dataclass(frozen=True, slots=True)
class SourceCatalog:
    """Every source this API will serve, and nothing else.

    The spreadsheet entries come from the reviewed contexts rather than from a
    directory listing, so the browser cannot open a sheet the system would never cite.
    """

    policy: Mapping[str, Path]
    scanned: Mapping[str, Path]
    sheets: Mapping[tuple[str, str], Path]
    tables: Mapping[str, SchemaContext]

    @classmethod
    def from_settings(cls, settings: Settings) -> SourceCatalog:
        sql_contexts = load_contexts(settings.sql_context_dir)
        sheet_contexts = load_contexts(settings.sheets_context_dir)

        sheets: dict[tuple[str, str], Path] = {}
        for context in sheet_contexts.values():
            if context.workbook is None or context.sheet is None:
                continue
            path = settings.spreadsheet_dir / context.workbook
            if path.is_file():
                sheets[(context.workbook, context.sheet)] = path

        return cls(
            policy=_pdfs(settings.policy_dir),
            scanned=_pdfs(settings.scanned_dir),
            sheets=sheets,
            tables={**sql_contexts, **sheet_contexts},
        )


def _document(corpus: Mapping[str, Path], name: str, kind: str) -> FileResponse:
    path = corpus.get(name)
    if path is None or not path.is_file():
        raise HTTPException(status_code=404, detail=f"No {kind} document named '{name}'")
    return FileResponse(
        path,
        media_type="application/pdf",
        filename=path.name,
        # Inline, or the browser hands the file to a download instead of showing it.
        content_disposition_type="inline",
    )


def build_router(catalog: SourceCatalog) -> APIRouter:
    router = APIRouter(prefix="/api/sources")

    @router.get("/policy/{document}")
    def policy_document(document: str) -> FileResponse:
        return _document(catalog.policy, document, "policy")

    @router.get("/scanned/{document}")
    def scanned_document(document: str) -> FileResponse:
        return _document(catalog.scanned, document, "scanned")

    @router.get("/spreadsheet/{workbook}/{sheet}")
    def spreadsheet(workbook: str, sheet: str) -> dict[str, object]:
        path = catalog.sheets.get((workbook, sheet))
        if path is None:
            raise HTTPException(
                status_code=404,
                detail=f"No reviewed sheet '{sheet}' in workbook '{workbook}'",
            )
        return read_sheet(path, sheet)

    @router.get("/sql/{table}")
    def sql_table(table: str) -> dict[str, object]:
        """What a SQL claim traces back to: the reviewed description of its table.

        No query runs. The rows a fresh query returned need not be the rows the
        answer used, and it would sit outside the validated plan path.
        """
        context = catalog.tables.get(table)
        if context is None:
            raise HTTPException(
                status_code=404, detail=f"No reviewed context for table '{table}'"
            )
        return context_detail(context)

    return router
