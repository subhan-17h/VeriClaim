"""Generate the deterministic semi-structured spreadsheet corpus."""

from __future__ import annotations

import io
import re
import zipfile
from collections import defaultdict
from collections.abc import Callable, Iterable, Sequence
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet

from vericlaim.config import get_settings
from vericlaim.corpus.catalog import ADJUSTERS, COVERAGE_PRODUCTS, REGIONS
from vericlaim.corpus.transactions import TransactionRows, generate_transactions
from vericlaim.sql.contexts import LINEAGE_COLUMN_NAMES, SchemaContext, load_contexts

FIXED_DOCUMENT_TIME = datetime(2026, 1, 1)
FIXED_CORE_TIME = b"2026-01-01T00:00:00Z"
FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
FIXED_ZIP_MODE = 0o600 << 16
DAY_QUANTUM = Decimal("0.01")

# Reviewed planning variances keep targets distinct from actuals while their scale
# follows the generated Q1 claim counts and incurred amounts as economics change.
CLAIMS_TARGET_VARIANCE_BASIS_POINTS = {
    1: 800,
    2: -700,
    3: 1200,
    4: -900,
    5: 600,
    6: -1200,
    7: 1000,
    8: -600,
    9: 1400,
}

INSPECTION_COMPLETION_BASIS_POINTS = {
    1: 6000,
    2: 7100,
    3: 8400,
    4: 7600,
    5: 6800,
    6: 7900,
    7: 8800,
    8: 8100,
    9: 7300,
}

RISK_CATEGORIES = (
    ("Modern reinforced concrete", "low", Decimal("0.05"), "Recent engineered construction."),
    ("Owner-occupied masonry", "medium", Decimal("0.10"), "Standard residential masonry."),
    ("Older masonry", "high", Decimal("0.20"), "Age and maintenance increase loss severity."),
    (
        "High-rise apartment",
        "medium",
        Decimal("0.12"),
        "Shared services affect escape-of-water risk.",
    ),
    (
        "Warehouse and light industrial",
        "high",
        Decimal("0.25"),
        "Contents and fire load vary by trade.",
    ),
    (
        "Flood-exposed property",
        "very high",
        Decimal("0.35"),
        "Loading requires current flood-zone review.",
    ),
)


def _save_deterministic(book: OpenpyxlWorkbook, path: Path) -> None:
    """Save an openpyxl workbook with byte-stable properties and ZIP metadata."""
    book.properties.created = FIXED_DOCUMENT_TIME
    book.properties.modified = FIXED_DOCUMENT_TIME
    book.properties.creator = "VeriClaim"
    book.properties.lastModifiedBy = "VeriClaim"

    source = io.BytesIO()
    book.save(source)
    source.seek(0)

    rewritten = io.BytesIO()
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(
        rewritten,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as outgoing:
        for name in sorted(incoming.namelist()):
            data = incoming.read(name)
            if name == "docProps/core.xml":
                data, replacements = re.subn(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>" + FIXED_CORE_TIME + rb"\g<2>",
                    data,
                )
                if replacements != 1:
                    raise ValueError("Could not pin docProps/core.xml modified timestamp")

            info = zipfile.ZipInfo(name, date_time=FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = FIXED_ZIP_MODE
            outgoing.writestr(info, data, compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(rewritten.getvalue())


def _data_columns(context: SchemaContext) -> tuple[str, ...]:
    return tuple(
        column.name
        for column in context.columns
        if column.name not in LINEAGE_COLUMN_NAMES
    )


def _new_book(context: SchemaContext) -> tuple[OpenpyxlWorkbook, Worksheet, Path]:
    if context.workbook is None or context.sheet is None:
        raise ValueError(f"{context.qualified} is not a workbook context")
    book = Workbook()
    sheet = book.active
    sheet.title = context.sheet
    return book, sheet, Path(context.workbook)


def _write_header(
    sheet: Worksheet, row: int, columns: Sequence[str], positions: Sequence[int]
) -> None:
    for column, position in zip(columns, positions, strict=True):
        cell = sheet.cell(row=row, column=position, value=column)
        cell.font = Font(bold=True)


def _write_rows(
    sheet: Worksheet,
    start_row: int,
    rows: Iterable[Sequence[object]],
    positions: Sequence[int],
) -> int:
    row_number = start_row
    for values in rows:
        for value, position in zip(values, positions, strict=True):
            sheet.cell(row=row_number, column=position, value=value)
        row_number += 1
    return row_number


def generate_loss_ratio_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 4, 5, 6)

    policy_by_id = {policy.policy_id: policy for policy in transactions.policies}
    premiums: defaultdict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0.00"))
    incurred: defaultdict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0.00"))
    for policy in transactions.policies:
        premiums[(policy.region_id, policy.product_id)] += policy.annual_premium_pkr
    for claim in transactions.claims:
        product_id = policy_by_id[claim.policy_id].product_id
        incurred[(claim.region_id, product_id)] += claim.incurred_amount_pkr

    rows = []
    for region in REGIONS:
        for product in COVERAGE_PRODUCTS:
            earned = premiums[(region.region_id, product.product_id)]
            claims = incurred[(region.region_id, product.product_id)]
            ratio = claims / earned
            rows.append((region.region_name, product.product_name, earned, claims, ratio))

    sheet["A1"] = "Loss Ratio Report"
    sheet.merge_cells("A1:F1")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    _write_header(sheet, 3, columns, positions)
    footer_row = _write_rows(sheet, 4, rows, positions)

    total_premium = sum((row[2] for row in rows), Decimal("0.00"))
    total_incurred = sum((row[3] for row in rows), Decimal("0.00"))
    footer = ("TOTAL", None, total_premium, total_incurred, total_incurred / total_premium)
    _write_rows(sheet, footer_row, (footer,), positions)
    for row in range(4, footer_row + 1):
        sheet.cell(row, 4).number_format = "PKR #,##0.00"
        sheet.cell(row, 5).number_format = "PKR #,##0.00"
        sheet.cell(row, 6).number_format = "0.00%"

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


def generate_adjuster_performance_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 3, 4)
    region_by_id = {region.region_id: region.region_name for region in REGIONS}

    assigned: defaultdict[int, int] = defaultdict(int)
    close_days: defaultdict[int, list[int]] = defaultdict(list)
    for claim in transactions.claims:
        if claim.adjuster_id is None:
            continue
        assigned[claim.adjuster_id] += 1
        if claim.status == "closed" and claim.closed_date is not None:
            close_days[claim.adjuster_id].append((claim.closed_date - claim.report_date).days)

    rows = []
    for adjuster in ADJUSTERS:
        days = close_days[adjuster.adjuster_id]
        average: object = (
            (Decimal(sum(days)) / Decimal(len(days))).quantize(DAY_QUANTUM, ROUND_HALF_UP)
            if days
            else "N/A"
        )
        if adjuster.adjuster_id == 6:
            average = "N/A"
        elif adjuster.adjuster_id == 14:
            average = "-"
        rows.append(
            (
                adjuster.adjuster_name,
                region_by_id[adjuster.region_id],
                assigned[adjuster.adjuster_id],
                average,
            )
        )

    _write_header(sheet, 1, columns, positions)
    _write_header(sheet, 2, columns, positions)
    _write_rows(sheet, 3, rows, positions)

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


def generate_claims_targets_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 3)

    q1_claims: defaultdict[int, int] = defaultdict(int)
    q1_incurred: defaultdict[int, Decimal] = defaultdict(lambda: Decimal("0.00"))
    for claim in transactions.claims:
        if claim.report_date.month > 3:
            continue
        q1_claims[claim.region_id] += 1
        q1_incurred[claim.region_id] += claim.incurred_amount_pkr

    rows = []
    for region in REGIONS:
        variance = CLAIMS_TARGET_VARIANCE_BASIS_POINTS[region.region_id]
        plan_factor = Decimal(10_000 + variance) / Decimal(10_000)
        target_claims = int(
            (Decimal(q1_claims[region.region_id]) * plan_factor).to_integral_value(
                rounding=ROUND_HALF_UP
            )
        )
        target_incurred = (q1_incurred[region.region_id] * plan_factor).quantize(
            DAY_QUANTUM, ROUND_HALF_UP
        )
        rows.append((region.region_name, target_claims, target_incurred))

    _write_header(sheet, 1, columns, positions)
    footer_row = _write_rows(sheet, 2, rows, positions)
    footer = (
        "TOTAL",
        sum(row[1] for row in rows),
        sum((row[2] for row in rows), Decimal("0.00")),
    )
    _write_rows(sheet, footer_row, (footer,), positions)
    for row in range(2, footer_row + 1):
        sheet.cell(row, 3).number_format = "PKR #,##0.00"

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


def generate_inspection_compliance_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 3, 4, 5)

    policies_by_region: defaultdict[int, int] = defaultdict(int)
    for policy in transactions.policies:
        policies_by_region[policy.region_id] += 1

    rows = []
    for region in REGIONS:
        # Roughly one planned inspection per six policies, rounded to a readable ten.
        scheduled = ((policies_by_region[region.region_id] + 30) // 60) * 10
        basis_points = INSPECTION_COMPLETION_BASIS_POINTS[region.region_id]
        completed = scheduled * basis_points // 10_000
        rate = Decimal(completed) / Decimal(scheduled)
        displayed_rate: object = "60%" if region.region_id == 1 else rate
        target_rate = Decimal("0.85") if region.region_id <= 6 else Decimal("0.90")
        rows.append(
            (region.region_name, scheduled, completed, displayed_rate, target_rate)
        )

    sheet["A1"] = "Regional Inspection Compliance - Q1"
    sheet.merge_cells("A1:E1")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    _write_header(sheet, 3, columns, positions)
    footer_row = _write_rows(sheet, 4, rows, positions)
    total_scheduled = sum(row[1] for row in rows)
    total_completed = sum(row[2] for row in rows)
    footer = (
        "TOTAL",
        total_scheduled,
        total_completed,
        Decimal(total_completed) / Decimal(total_scheduled),
        None,
    )
    _write_rows(sheet, footer_row, (footer,), positions)
    for row in range(4, footer_row + 1):
        sheet.cell(row, 4).number_format = "0.00%"
        sheet.cell(row, 5).number_format = "0.00%"

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


def generate_renewals_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 4, 5, 6)

    due: defaultdict[tuple[int, int], int] = defaultdict(int)
    renewed: defaultdict[tuple[int, int], int] = defaultdict(int)
    for policy in transactions.policies:
        # The first annual anniversary of a Q1 2025 inception is due in Q1 2026.
        if policy.inception_date.month > 3:
            continue
        key = (policy.region_id, policy.product_id)
        due[key] += 1
        if policy.status == "active":
            renewed[key] += 1

    rows = []
    for region in REGIONS:
        for product in COVERAGE_PRODUCTS:
            key = (region.region_id, product.product_id)
            rate = Decimal(renewed[key]) / Decimal(due[key])
            rows.append(
                (
                    region.region_name,
                    product.product_name,
                    due[key],
                    renewed[key],
                    rate,
                )
            )

    _write_header(sheet, 1, columns, positions)
    footer_row = _write_rows(sheet, 2, rows, positions)
    total_due = sum(row[2] for row in rows)
    total_renewed = sum(row[3] for row in rows)
    footer = (
        "TOTAL",
        None,
        total_due,
        total_renewed,
        Decimal(total_renewed) / Decimal(total_due),
    )
    _write_rows(sheet, footer_row, (footer,), positions)
    for row in range(2, footer_row + 1):
        sheet.cell(row, 6).number_format = "0.00%"

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


def generate_risk_categories_workbook(
    context: SchemaContext, transactions: TransactionRows, output_dir: Path
) -> Path:
    del transactions
    book, sheet, filename = _new_book(context)
    columns = _data_columns(context)
    positions = (1, 2, 3, 4)

    _write_header(sheet, 1, columns, positions)
    _write_rows(sheet, 2, RISK_CATEGORIES, positions)
    for row in range(2, 2 + len(RISK_CATEGORIES)):
        sheet.cell(row, 3).number_format = "0%"

    path = output_dir / filename
    _save_deterministic(book, path)
    return path


WORKBOOK_GENERATORS: dict[
    str, Callable[[SchemaContext, TransactionRows, Path], Path]
] = {
    "adjuster_performance__performance": generate_adjuster_performance_workbook,
    "claims_targets_q1__targets": generate_claims_targets_workbook,
    "loss_ratio_report__loss_ratio": generate_loss_ratio_workbook,
    "regional_inspection_compliance_q1__compliance": generate_inspection_compliance_workbook,
    "renewals_q1__renewals": generate_renewals_workbook,
    "risk_categories__categories": generate_risk_categories_workbook,
}


def generate_spreadsheet_corpus(
    output_dir: Path | None = None, *, seed: int = 42
) -> list[Path]:
    """Write exactly the workbooks declared by the reviewed sheet contexts."""
    settings = get_settings()
    destination = output_dir or settings.spreadsheet_dir
    contexts = load_contexts(settings.sheets_context_dir)
    by_table = {context.table: context for context in contexts.values()}
    if by_table.keys() != WORKBOOK_GENERATORS.keys():
        missing = sorted(by_table.keys() - WORKBOOK_GENERATORS.keys())
        unexpected = sorted(WORKBOOK_GENERATORS.keys() - by_table.keys())
        raise ValueError(
            f"Spreadsheet generators do not match contexts: missing={missing}, "
            f"unexpected={unexpected}"
        )

    transactions = generate_transactions(seed)
    return [
        WORKBOOK_GENERATORS[table](by_table[table], transactions, destination)
        for table in sorted(by_table)
    ]


def main() -> None:
    for path in generate_spreadsheet_corpus():
        print(path.name)


if __name__ == "__main__":
    main()
